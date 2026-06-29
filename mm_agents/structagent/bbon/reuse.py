"""Reconstruct bBoN RolloutArtifacts from EXISTING run dumps — no re-run, no VM.

Builds the OFFLINE judge frozen set cheaply from the ~8k rollouts already on
disk (tasks repeat 20-30x across configs/dates). Per task dir we read:

  result.txt                                   -> real eval score (0/1)
  keynode_debug/step_NNN_keynode_input.json    -> per-step outcomes + action
  keynode_debug/step_NNN_keynode_verdicts.json -> per-outcome deterministic_trace
  planner_debug/step_NNNN_*.html               -> screenshots (embedded png; take max)
  traj.jsonl                                   -> per-step response (for <decision>)

Synthetic KeyNodes come from outcome state transitions (pending->verified =
satisfied; verified->reverted = invalidated), with the step's deterministic_trace
as evidence. Keeps the facts-not-verdicts rule (narrative.py): render the probe
trace, never "satisfied". Production bBoN still runs N live rollouts.
"""
from __future__ import annotations

import base64
import glob
import json
import os
import re
from collections import defaultdict
from types import SimpleNamespace
from typing import Any, Dict, List, Optional, Tuple

from mm_agents.structagent.ledger.core.ledger import Outcome
from mm_agents.structagent.ledger.core.timeline import (
    StepRecord, KeyNode, KN_OUTCOME_SATISFIED, KN_OUTCOME_INVALIDATED,
)
from mm_agents.structagent.bbon.narrative import RolloutArtifacts

_TASK_ID_RE = re.compile(r"/([a-f0-9]{8}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{12})/")
_PNG_RE = re.compile(r"data:image/png;base64,([A-Za-z0-9+/=]{200,})")
_DECISION_RE = re.compile(r"<decision>\s*([A-Za-z_]+)\s*</decision>", re.IGNORECASE)
_PLANNER_STEP_RE = re.compile(r"step_(\d+)_")


def task_id_of(path: str) -> Optional[str]:
    m = _TASK_ID_RE.search(path if path.endswith("/") else path + "/")
    return m.group(1) if m else None


def load_score(task_dir: str) -> Optional[float]:
    p = os.path.join(task_dir, "result.txt")
    if not os.path.exists(p):
        return None
    try:
        return float(open(p).read().strip().split()[0])
    except Exception:
        return None


def _trace_summary(trace: Any) -> str:
    """Objective one-liner for a deterministic_trace (probe + raw result), no verdict."""
    if not isinstance(trace, dict):
        return str(trace)[:200]
    kind = trace.get("kind", "probe")
    bits: List[str] = []
    cmd = trace.get("command")
    if isinstance(cmd, list):
        bits.append("`" + " ".join(map(str, cmd))[:140] + "`")
    for k in ("file_path", "patterns", "returncode", "exit_code", "stdout", "matched", "value"):
        if trace.get(k) not in (None, "", [], {}):
            bits.append(f"{k}={str(trace[k])[:80]}")
    return f"probe[{kind}] " + " ".join(bits) if bits else f"probe[{kind}]"


def _load_steps(task_dir: str) -> List[dict]:
    """keynode_debug -> per-step dicts {step_idx, action, states{id:state},
    traces{id:summary}}, ordered by step_idx."""
    kd = os.path.join(task_dir, "keynode_debug")
    inputs: Dict[int, dict] = {}
    for p in glob.glob(os.path.join(kd, "*_keynode_input.json")):
        try:
            o = json.load(open(p))
            inputs[int(o["step_idx"])] = o
        except Exception:
            continue
    verdicts: Dict[int, dict] = {}
    for p in glob.glob(os.path.join(kd, "*_keynode_verdicts.json")):
        try:
            o = json.load(open(p))
            verdicts[int(o["step_idx"])] = o
        except Exception:
            continue
    steps = []
    for si in sorted(inputs):
        inp = inputs[si]
        states = {o["id"]: o.get("current_state", "pending") for o in inp.get("outcomes", [])}
        traces = {}
        for po in (verdicts.get(si, {}) or {}).get("per_outcome", []):
            t = po.get("deterministic_trace")
            if t:
                traces[po.get("outcome_id")] = _trace_summary(t)
        steps.append({"step_idx": si,
                      "action": inp.get("actor_action_summary", "") or "",
                      "states": states, "traces": traces})
    if steps:
        return steps
    # Boundary-mode / keynode-less runs emit no per-step *_keynode_input.json,
    # so rebuild a skeleton from traj.jsonl (no per-step outcome states; judge
    # leans on screenshots + actions + final).
    return _load_steps_from_traj(task_dir)


def _load_steps_from_traj(task_dir: str) -> List[dict]:
    """Step skeleton from traj.jsonl ({step_idx, action, states{}, traces{}}),
    for when keynode_debug has no per-step input dumps."""
    p = os.path.join(task_dir, "traj.jsonl")
    if not os.path.exists(p):
        return []
    out: List[dict] = []
    for line in open(p):
        line = line.strip()
        if not line:
            continue
        try:
            r = json.loads(line)
        except Exception:
            continue
        si = r.get("step_num")
        if si is None:
            continue
        resp = r.get("response") or ""
        if "Action:" in resp:
            action = resp.split("Action:", 1)[1].strip()[:200]
        else:
            action = (r.get("action") or "")[:200]
        out.append({"step_idx": int(si), "action": action,
                    "states": {}, "traces": {}})
    return out


def _planner_html_index(task_dir: str) -> Dict[int, str]:
    """{step_idx: planner_debug html path} from filenames (no read)."""
    idx: Dict[int, str] = {}
    for p in glob.glob(os.path.join(task_dir, "planner_debug", "step_*.html")):
        m = _PLANNER_STEP_RE.search(os.path.basename(p))
        if m:
            idx.setdefault(int(m.group(1)), p)
    return idx


def _load_screenshots(task_dir: str, needed: Optional[set] = None) -> Dict[int, str]:
    """planner_debug html -> {step_idx: largest embedded png b64}. Reads only the
    htmls for ``needed`` steps — each is multi-MB, so targeted loading avoids the
    whole 100-step trajectory. ``needed=None`` reads all (slow; avoid)."""
    idx = _planner_html_index(task_dir)
    items = idx.items() if needed is None else ((s, idx[s]) for s in needed if s in idx)
    out: Dict[int, str] = {}
    for s, p in items:
        try:
            imgs = _PNG_RE.findall(open(p, encoding="utf-8", errors="ignore").read())
        except Exception:
            continue
        if imgs:
            out[s] = max(imgs, key=len)   # largest = full screenshot
    return out


def _load_decisions(task_dir: str) -> Dict[int, str]:
    """traj.jsonl -> {step_idx: planner decision}. step_num is 1-based."""
    out: Dict[int, str] = {}
    p = os.path.join(task_dir, "traj.jsonl")
    if not os.path.exists(p):
        return out
    for line in open(p, encoding="utf-8", errors="ignore"):
        try:
            r = json.loads(line)
        except Exception:
            continue
        m = _DECISION_RE.search(r.get("response", "") or "")
        if m and r.get("step_num") is not None:
            out[int(r["step_num"]) - 1] = m.group(1).upper()
    return out


def _load_final_state_facts(task_dir: str) -> str:
    """Final-state facts from the LAST perceiver snapshot: visible controls
    (label/role/state/value) + transition summary. Excludes the perceiver's own
    subgoal verdict — facts, not another model's judgment."""
    snaps = sorted(glob.glob(os.path.join(task_dir, "perceiver_debug",
                                          "*_perceiver_snapshot.json")))
    if not snaps:
        return ""
    try:
        d = json.load(open(snaps[-1]))
    except Exception:
        return ""
    parts: List[str] = []
    ts = d.get("transition_summary")
    if ts:
        parts.append(f"transition: {str(ts)[:300]}")
    ctrls = d.get("relevant_controls") or []
    rows = []
    for c in (ctrls if isinstance(ctrls, list) else [])[:12]:
        if not isinstance(c, dict):
            continue
        lbl = str(c.get("label", ""))[:50]
        role = c.get("role", "")
        st = c.get("state") or []
        val = c.get("value")
        row = f"{role}: {lbl}"
        if st:
            row += f" [{','.join(str(x) for x in st)}]"
        if val not in (None, ""):
            row += f" = {val}"
        rows.append(row.strip())
    if rows:
        parts.append("visible controls:\n  " + "\n  ".join(rows))
    return "\n".join(parts)


def _load_step_a11y(task_dir: str) -> Dict[int, str]:
    """{step_num: a11y_filtered_text} from perceiver_debug (URL + element text +
    state). Paired with each frame so the judge reads page/filter state instead
    of guessing from pixels."""
    out: Dict[int, str] = {}
    for p in glob.glob(os.path.join(task_dir, "perceiver_debug",
                                    "*_perceiver_input.json")):
        m = re.search(r"step_(\d+)_perceiver_input", os.path.basename(p))
        if not m:
            continue
        try:
            txt = json.load(open(p)).get("a11y_filtered_text", "") or ""
        except Exception:
            continue
        if txt:
            out[int(m.group(1))] = txt
    return out


def from_existing_run(task_dir: str, *, with_screenshots: bool = True) -> Optional[RolloutArtifacts]:
    """Rebuild RolloutArtifacts from one task dir, or None if it has no steps to
    reconstruct from. Synthetic KeyNodes come from outcome state transitions;
    evidence = the deterministic_trace."""
    steps = _load_steps(task_dir)
    if not steps:
        return None
    decisions = _load_decisions(task_dir)
    step_a11y = _load_step_a11y(task_dir)

    timeline: List[StepRecord] = []
    prev_states: Dict[str, str] = {}
    for st in steps:
        si = st["step_idx"]
        kns: List[KeyNode] = []
        for oid, state in st["states"].items():
            prev = prev_states.get(oid)
            if prev is not None and state != prev:
                if state == "verified":
                    kind = KN_OUTCOME_SATISFIED
                elif state == "reverted":
                    kind = KN_OUTCOME_INVALIDATED
                else:
                    kind = None
                if kind:
                    kns.append(KeyNode(
                        kind=kind, target=oid,
                        evidence=st["traces"].get(oid, "") or f"({oid} state -> {state})",
                        confidence="high", detected_at_step=si, source="deterministic"))
        prev_states = dict(st["states"])
        timeline.append(StepRecord(
            step_idx=si, screenshot_b64=None,
            a11y=step_a11y.get(si, ""),
            actor_action_summary=st["action"],
            planner_decision=decisions.get(si), key_nodes=kns))

    # Targeted screenshot load: turning-point candidates + first/last anchors
    # (each html is multi-MB; reading all 100 is the slow path).
    final_shot = ""
    shots: Dict[int, str] = {}
    if with_screenshots and timeline:
        first_s, last_s = timeline[0].step_idx, timeline[-1].step_idx
        # Always include the last few steps' frames so boundary-mode runs (no
        # per-step keynodes) still get decisive tail frames, not just first/last.
        from mm_agents.structagent.bbon.narrative import DEFAULT_MAX_FRAMES as _MF
        needed = {first_s, last_s} | {s.step_idx for s in timeline[-_MF:]} | {
            s.step_idx for s in timeline
            if s.key_nodes or (s.planner_decision or "").upper() in ("REPLAN", "DONE")}
        shots = _load_screenshots(task_dir, needed)
        for s in timeline:
            if s.step_idx in shots:
                s.screenshot_b64 = shots[s.step_idx]
        final_shot = shots.get(last_s, "")

    # Outcomes with final states from the last step's snapshot.
    final_states = steps[-1]["states"]
    outcomes = [Outcome(id=o["id"], description=o.get("description", ""),
                        evidence_hint=o.get("evidence_hint", ""),
                        state=final_states.get(o["id"], "pending"))
                for o in _last_outcome_defs(task_dir)]
    ledger = SimpleNamespace(required_outcomes=outcomes)

    # Last step's perceiver a11y (closest to terminal state).
    final_a11y = step_a11y[max(step_a11y)] if step_a11y else ""
    # Infeasible signal: agent terminated with FAIL / impossible.
    emitted_fail = any(
        kw in (s.actor_action_summary or "")
        for s in timeline for kw in ("FAIL", "impossible", "Impossible", "IMPOSSIBLE"))
    art = RolloutArtifacts(
        task=_load_task_instruction(task_dir),
        ledger=ledger, timeline=timeline,
        final_obs={"screenshot": final_shot, "a11y_text": final_a11y},
        final_doc_structure=_load_final_doc_structure(task_dir),
        emitted_fail=emitted_fail,
    )
    return art


def _load_final_doc_structure(task_dir: str) -> str:
    """Parse the FINAL output file(s) under ``eval_files/`` — the SAME files the
    OSWorld checker reads — into compact text (cell values + formulas + number
    formats). This checker-level state is what screenshots and AT-SPI a11y lack.
    Uses the real exported file, not the SheetPerceiver probe (which may be stale).
    Empty when nothing parseable (judge falls back to screenshots)."""
    ed = os.path.join(task_dir, "eval_files")
    if not os.path.isdir(ed):
        return ""
    blocks: List[str] = []
    for f in sorted(os.listdir(ed)):
        p = os.path.join(ed, f)
        low = f.lower()
        if low.endswith((".xlsx", ".xlsm", ".xls")):
            blocks.append(f"FILE {f}:\n" + _render_xlsx(p))
        elif low.endswith(".csv"):
            blocks.append(f"FILE {f}:\n" + _render_csv(p))
        elif low.endswith(".docx"):
            blocks.append(f"FILE {f}:\n" + _render_docx(p))
        elif low.endswith(".odt"):
            blocks.append(f"FILE {f}:\n" + _render_odt(p))
        elif low.endswith((".json", ".py", ".txt", ".code-workspace",
                           ".md", ".xml", ".yaml", ".yml", ".js", ".ts",
                           ".html", ".css", ".sh", ".cfg", ".ini")):
            blocks.append(f"FILE {f}:\n" + _render_textfile(p))
        # pdf / images / other non-tabular: skip (judge keeps screenshots).
        if sum(len(b) for b in blocks) > 3500:
            break
    return "\n\n".join(blocks)[:3500]


def _render_xlsx(path: str, *, max_rows: int = 40, max_cols: int = 20) -> str:
    """openpyxl → text: per sheet, used range with formulas + cached values
    (``=SUM(...)->4852``) and non-default number formats. Loads both the checker's
    formula view (default) and data_only for the computed result."""
    try:
        import openpyxl
        wf = openpyxl.load_workbook(path, data_only=False)   # formulas
        wv = openpyxl.load_workbook(path, data_only=True)    # cached values
    except Exception as e:
        return f"  [xlsx parse failed: {str(e)[:80]}]"
    out: List[str] = []
    for name in wf.sheetnames:
        sf, sv = wf[name], wv[name]
        mr, mc = min(sf.max_row, max_rows), min(sf.max_column, max_cols)
        if mr < 1 or mc < 1:
            continue
        out.append(f"── Sheet[{name}] used={sf.dimensions} ──")
        fmts = []
        for r in range(1, mr + 1):
            cells, allnone = [], True
            for c in range(1, mc + 1):
                fc = sf.cell(r, c)
                f, v = fc.value, sv.cell(r, c).value
                if f is None and v is None:
                    cells.append("")
                    continue
                allnone = False
                if isinstance(f, str) and f.startswith("=") and v is not None and v != f:
                    cells.append(f"{f}->{v}")
                else:
                    cells.append(str(f if f is not None else v))
                nf = (fc.number_format or "General")
                if nf != "General":
                    fmts.append(f"{fc.coordinate}={nf}")
            if not allnone:
                out.append(f"  [{r:>3}] " + " | ".join(cells).rstrip(" |"))
        if sf.max_row > max_rows:
            out.append(f"  ... ({sf.max_row - max_rows} more rows)")
        if fmts:
            out.append("  number-formats: " + "; ".join(fmts[:20]))
        # Charts/sparklines: many calc tasks ARE "make a chart", invisible in cell
        # values, so surface them explicitly (type + data range).
        charts = getattr(sf, "_charts", []) or []
        if charts:
            crows = []
            for ch in charts[:6]:
                refs = []
                try:
                    for se in (ch.series or [])[:4]:
                        nr = getattr(getattr(se, "val", None), "numRef", None)
                        if nr and getattr(nr, "f", None):
                            refs.append(str(nr.f))
                except Exception:
                    pass
                crows.append(type(ch).__name__
                             + (f"(data={'; '.join(refs)})" if refs else ""))
            out.append(f"  charts ({len(charts)}): " + " | ".join(crows))
        try:
            sgs = getattr(sf, "sparkline_groups", None) or []
            if sgs:
                out.append(f"  sparkline-groups: {len(sgs)}")
        except Exception:
            pass
    return "\n".join(out) if out else "  (empty workbook)"


def _render_csv(path: str, *, max_rows: int = 40) -> str:
    import csv as _csv
    try:
        rows = list(_csv.reader(open(path, newline="", errors="ignore")))
    except Exception as e:
        return f"  [csv parse failed: {str(e)[:80]}]"
    out = [f"  [{i+1:>3}] " + " | ".join(r) for i, r in enumerate(rows[:max_rows])]
    if len(rows) > max_rows:
        out.append(f"  ... ({len(rows) - max_rows} more rows)")
    return "\n".join(out) if out else "  (empty csv)"


def _render_docx(path: str, *, max_paras: int = 80) -> str:
    """.docx paragraph text + per-run formatting (font/bold/italic/underline/size)
    — the checker-level document state. Writer tasks are mostly about formatting,
    invisible in screenshots, so the flags ride next to the text."""
    try:
        from docx import Document
    except Exception as e:
        return f"  [python-docx unavailable: {str(e)[:60]}]"
    try:
        doc = Document(path)
    except Exception as e:
        return f"  [docx parse failed: {str(e)[:80]}]"
    lines: List[str] = []
    for para in doc.paragraphs[:max_paras]:
        txt = (para.text or "").strip()
        if not txt:
            continue
        fmts = set()
        for r in para.runs:
            if r.bold:
                fmts.add("bold")
            if r.italic:
                fmts.add("italic")
            if r.underline:
                fmts.add("underline")
            nm = getattr(r.font, "name", None)
            if nm:
                fmts.add(f"font={nm}")
            sz = getattr(r.font, "size", None)
            if sz is not None:
                try:
                    fmts.add(f"size={sz.pt:g}")
                except Exception:
                    pass
            try:
                col = getattr(r.font, "color", None)
                if col is not None and getattr(col, "rgb", None) is not None:
                    fmts.add(f"color=#{col.rgb}")
            except Exception:
                pass
        # Paragraph alignment — common writer checker target, invisible in screenshots.
        al = getattr(para, "alignment", None)
        if al is not None:
            _almap = {0: "left", 1: "center", 2: "right", 3: "justify",
                      4: "distribute"}
            try:
                fmts.add(f"align={_almap.get(int(al), str(al))}")
            except Exception:
                fmts.add(f"align={str(al)}")
        pf = getattr(para, "paragraph_format", None)
        if pf is not None:
            ls = getattr(pf, "line_spacing", None)
            if ls is not None:
                try:
                    fmts.add(f"linespacing={float(ls):g}")
                except Exception:
                    pass
            for _iname, _attr in (("indent", "left_indent"),
                                  ("firstline", "first_line_indent")):
                _v = getattr(pf, _attr, None)
                if _v is not None:
                    try:
                        fmts.add(f"{_iname}={_v.pt:g}pt")
                    except Exception:
                        pass
        sty = getattr(getattr(para, "style", None), "name", None)
        tag = f"  [{', '.join(sorted(fmts))}]" if fmts else ""
        styp = f" <{sty}>" if sty and sty != "Normal" else ""
        lines.append(f"{txt[:140]}{styp}{tag}")
    return "\n".join(lines) if lines else "  (empty document)"


def _render_textfile(path: str, *, max_chars: int = 1800) -> str:
    """Plain-text output file (json/py/txt/code-workspace/...) verbatim — the
    checker-level content for vs_code and other text-output tasks."""
    try:
        txt = open(path, encoding="utf-8", errors="replace").read()
    except Exception as e:
        return f"  [read failed: {str(e)[:80]}]"
    txt = txt.strip()
    if len(txt) > max_chars:
        txt = txt[:max_chars] + f"\n  ... ({len(txt) - max_chars} more chars)"
    return txt or "  (empty file)"


def _render_odt(path: str, *, max_paras: int = 80) -> str:
    """.odt paragraph text + style name via odfpy — the odt counterpart of
    _render_docx. ODF stores formatting by style reference, so we surface text +
    the paragraph style name (carries alignment/heading for most checker tasks)."""
    try:
        from odf.opendocument import load
        from odf import text as _odftext, teletype
    except Exception as e:
        return f"  [odfpy unavailable: {str(e)[:60]}]"
    try:
        doc = load(path)
        paras = doc.getElementsByType(_odftext.P)
    except Exception as e:
        return f"  [odt parse failed: {str(e)[:80]}]"
    lines: List[str] = []
    for p in paras[:max_paras]:
        try:
            t = teletype.extractText(p).strip()
        except Exception:
            t = ""
        if not t:
            continue
        sn = p.getAttribute("stylename") or ""
        lines.append(f"{t[:140]}" + (f"  <{sn}>" if sn else ""))
    return "\n".join(lines) if lines else "  (empty odt)"


def _last_outcome_defs(task_dir: str) -> List[dict]:
    """Outcome id/description/evidence_hint defs (stable across steps)."""
    ins = sorted(glob.glob(os.path.join(task_dir, "keynode_debug", "*_keynode_input.json")))
    if not ins:
        return []
    try:
        return json.load(open(ins[-1])).get("outcomes", [])
    except Exception:
        return []


def _load_task_instruction(task_dir: str) -> str:
    """Authoritative task instruction from the task config json. Do NOT regex
    'Task:'/'Instruction:' out of init_ledger_prompt — that prompt embeds a few-shot
    EXAMPLE the regex hit first, mislabeling every task with the example and feeding
    the judge the wrong goal (tanked calc judge_acc; polluted chrome/gimp too)."""
    tid = task_id_of(task_dir) or os.path.basename(task_dir)
    for c in glob.glob(os.path.join("evaluation_examples", "examples", "*",
                                    f"{tid}.json")):
        try:
            instr = (json.load(open(c)).get("instruction") or "").strip()
            if instr:
                return instr[:500]
        except Exception:
            pass
    return f"(task {tid})"


def scan_scores(run_roots: List[str], *, domains: Optional[List[str]] = None,
                min_rollouts: int = 2) -> Dict[str, List[Tuple[str, int]]]:
    """Cheap first pass: group task dirs by id, read result.txt scores only (no
    artifact parsing / screenshots). Returns {task_id: [(dir, score01), ...]} for
    tasks with >= min_rollouts. Used to compute recoverable%/oracle/single for free
    and pick which dirs are worth the expensive screenshot parse."""
    by_task: Dict[str, List[Tuple[str, int]]] = defaultdict(list)
    for root in run_roots:
        for rt in glob.glob(os.path.join(root, "**", "result.txt"), recursive=True):
            d = os.path.dirname(rt)
            tid = task_id_of(d)
            if not tid:
                continue
            if domains and not any(f"/{dom}/" in d + "/" for dom in domains):
                continue
            sc = load_score(d)
            if sc is not None:
                by_task[tid].append((d, 1 if sc >= 0.999 else 0))
    return {t: v for t, v in by_task.items() if len(v) >= min_rollouts}


def build_frozen_set(run_roots: List[str], *, domains: Optional[List[str]] = None,
                     min_rollouts: int = 2, max_per_task: Optional[int] = None,
                     with_screenshots: bool = True) -> List[List[Tuple[RolloutArtifacts, int]]]:
    """Group task dirs by id, parse each rollout, keep tasks with >= min_rollouts.
    Returns the frozen_set for eval_judge (score = 1 if success>=0.999 else 0).
    ``with_screenshots=False`` is fast (recoverable%-only / text narratives)."""
    by_task: Dict[str, List[str]] = defaultdict(list)
    for root in run_roots:
        for rt in glob.glob(os.path.join(root, "**", "result.txt"), recursive=True):
            d = os.path.dirname(rt)
            tid = task_id_of(d)
            if not tid:
                continue
            if domains and not any(f"/{dom}/" in d for dom in domains):
                continue
            by_task[tid].append(d)

    frozen: List[List[Tuple[RolloutArtifacts, int]]] = []
    for tid, dirs in by_task.items():
        if len(dirs) < min_rollouts:
            continue
        rollouts: List[Tuple[RolloutArtifacts, int]] = []
        for d in (dirs[:max_per_task] if max_per_task else dirs):
            sc = load_score(d)
            if sc is None:
                continue
            art = from_existing_run(d, with_screenshots=with_screenshots)
            if art is not None:
                rollouts.append((art, 1 if sc >= 0.999 else 0))
        if len(rollouts) >= min_rollouts:
            frozen.append(rollouts)
    return frozen
