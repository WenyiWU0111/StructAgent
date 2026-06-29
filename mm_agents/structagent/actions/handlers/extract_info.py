"""Host-side handler for the ``extract_info`` actor action.

Reads N files from the VM via ``controller.get_file``, dispatches each to
the planner LLM (vision-capable) with a structured-output prompt, and
returns an aggregate dict the agent stores on ``self._last_extract_output``
for rendering into the next planner prompt.

Sequential by design — one file at a time, one LLM call per file. The
user explicitly chose sequential over concurrent to avoid race conditions
in the inference server and keep failure modes simple to reason about.
"""
from __future__ import annotations

import base64
import io
import json
import logging
import os
import re
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger("mm_agents.extract_info")


# Per-file LLM call timeout — chosen so the whole extract_info action
# (≤ 8 files in the typical receipts/papers tasks) stays under a few
# minutes even in the worst case.
PER_FILE_TIMEOUT_S = 60
# Cap on N files per extract_info call. Larger collections should be
# chunked across multiple actions so the planner gets intermediate
# results sooner. Hard cap to bound worst-case latency.
MAX_FILES_PER_CALL = 12
# Max bytes pulled per file (the planner LLM context window dictates
# this; bigger files get tail-truncated for text, downscaled for images).
MAX_FILE_BYTES = 8 * 1024 * 1024     # 8 MiB
# Truncation budget for raw text excerpt returned to the agent (kept
# small — planner only needs evidence the extraction is grounded).
RAW_EXCERPT_CHARS = 600


_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp"}
_PDF_EXTS = {".pdf"}
_OFFICE_TEXT_EXTS = {".txt", ".md", ".csv", ".log", ".json", ".yaml",
                     ".yml", ".xml", ".html", ".htm", ".eml"}
_DOCX_EXTS = {".docx"}
_XLSX_EXTS = {".xlsx", ".ods"}


def _ext_of(path: str) -> str:
    return os.path.splitext(path)[1].lower()


def _fetch_file(controller, path: str) -> Optional[bytes]:
    """Pull a file from the VM. Returns None when missing / oversized."""
    try:
        data = controller.get_file(path)
    except Exception as e:
        logger.warning("[extract_info] get_file(%s) raised: %s", path, e)
        return None
    if data is None:
        return None
    if len(data) > MAX_FILE_BYTES:
        logger.warning(
            "[extract_info] %s is %d bytes (cap %d) — truncating tail",
            path, len(data), MAX_FILE_BYTES,
        )
        data = data[:MAX_FILE_BYTES]
    return data


def _bytes_to_data_url(data: bytes, mime: str) -> str:
    return f"data:{mime};base64,{base64.b64encode(data).decode('ascii')}"


def _maybe_downscale_image(data: bytes, ext: str) -> bytes:
    """Downscale very large images so the VLM doesn't blow up context.
    Keep aspect ratio, long edge ≤ 1600px. Returns original bytes when
    PIL unavailable or downscale not needed."""
    try:
        from PIL import Image
    except ImportError:
        return data
    try:
        img = Image.open(io.BytesIO(data))
        long_edge = max(img.size)
        if long_edge <= 1600:
            return data
        ratio = 1600 / long_edge
        new_size = (int(img.size[0] * ratio), int(img.size[1] * ratio))
        img = img.resize(new_size, Image.LANCZOS)
        buf = io.BytesIO()
        fmt = "PNG" if ext in (".png", ".webp") else "JPEG"
        img.convert("RGB").save(buf, format=fmt, quality=85)
        return buf.getvalue()
    except Exception as e:
        logger.warning("[extract_info] downscale failed for %s ext: %s",
                       ext, e)
        return data


def _extract_pdf_text(data: bytes) -> Tuple[Optional[str], Optional[bytes]]:
    """Try text extraction first. Returns (text, None) on success.
    On failure or empty text, returns (None, first_page_image_bytes)
    for VLM fallback. (None, None) when both fail."""
    text = ""
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(data))
        text_parts = []
        for page in reader.pages[:10]:    # cap pages for context
            try:
                text_parts.append(page.extract_text() or "")
            except Exception:
                continue
        text = "\n".join(text_parts).strip()
    except Exception as e:
        logger.warning("[extract_info] pypdf failed: %s", e)
    if text:
        return text, None
    # Fall back to first-page render
    try:
        from pdf2image import convert_from_bytes
        images = convert_from_bytes(data, first_page=1, last_page=1, dpi=120)
        if images:
            buf = io.BytesIO()
            images[0].convert("RGB").save(buf, format="JPEG", quality=85)
            return None, buf.getvalue()
    except Exception as e:
        logger.warning("[extract_info] pdf2image fallback failed: %s", e)
    return None, None


def _extract_docx_text(data: bytes) -> str:
    try:
        from docx import Document
        doc = Document(io.BytesIO(data))
        return "\n".join(p.text for p in doc.paragraphs)
    except Exception as e:
        logger.warning("[extract_info] docx parse failed: %s", e)
        return ""


def _extract_xlsx_summary(data: bytes) -> str:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        lines = []
        for ws in wb.worksheets[:3]:
            lines.append(f"Sheet '{ws.title}':")
            for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if r_idx >= 20:
                    lines.append("  ...")
                    break
                lines.append("  " + " | ".join(str(c) if c is not None else "" for c in row[:12]))
        return "\n".join(lines)
    except Exception as e:
        logger.warning("[extract_info] xlsx parse failed: %s", e)
        return ""


def _build_llm_call(
    *, query: str, schema: Dict[str, Any], path: str,
    text_content: Optional[str], image_bytes: Optional[bytes],
    image_mime: str = "image/jpeg",
) -> List[Dict[str, Any]]:
    """Construct OpenAI-style chat messages for ONE file extraction.

    Returns the messages list (caller calls the LLM). The prompt is
    deliberately short: file path + query + schema + content. Output
    constraint: ONE JSON object literally matching `schema` keys (or a
    documented {"_skip": true, "reason": "..."} when the file doesn't
    carry the requested info).
    """
    sys = (
        "You extract STRUCTURED INFORMATION from a single file's content "
        "(image or text). Output STRICTLY one JSON object — no prose "
        "before or after, no markdown fences. The JSON must match the "
        "REQUESTED SCHEMA below.\n\n"
        "Rules:\n"
        "1. Every key in the schema MUST appear in your output.\n"
        "2. If a value is not present / not legible in the file, set "
        "that key to null (do NOT invent).\n"
        "3. Quote values VERBATIM where they appear (don't paraphrase). "
        "Dates: prefer ISO 'YYYY-MM-DD'.\n"
        "4. When the file is clearly NOT relevant to the query (e.g. an "
        "unrelated image), output {\"_skip\": true, \"reason\": "
        "\"<one-line>\"} instead of all nulls."
    )
    schema_block = (
        f"REQUESTED SCHEMA (output keys + types):\n{json.dumps(schema, indent=2)}"
        if schema else
        "OUTPUT SCHEMA: free-form JSON; pick keys that best capture the requested info."
    )
    user_text_parts = [
        f"File: {os.path.basename(path)}",
        f"Query: {query}",
        schema_block,
    ]
    if text_content is not None:
        if len(text_content) > 8000:
            text_content = text_content[:4000] + "\n\n[... truncated ...]\n\n" + text_content[-3000:]
        user_text_parts.append(f"--- FILE CONTENT ---\n{text_content}\n--- END ---")
    user_content: List[Dict[str, Any]] = [
        {"type": "text", "text": "\n\n".join(user_text_parts)},
    ]
    if image_bytes is not None:
        user_content.append({
            "type": "image_url",
            "image_url": {"url": _bytes_to_data_url(image_bytes, image_mime)},
        })
    return [
        {"role": "system", "content": sys},
        {"role": "user", "content": user_content},
    ]


_JSON_OBJECT_RE = re.compile(r"\{[\s\S]*\}")


def _parse_json_response(raw: str) -> Optional[Dict[str, Any]]:
    """Extract a single JSON object from the LLM response. Tolerates
    optional markdown fences and prose around the object."""
    if not raw:
        return None
    s = raw.strip()
    # Strip ```json fences
    m = re.match(r"^```(?:json)?\s*\n([\s\S]*?)\n```\s*$", s)
    if m:
        s = m.group(1).strip()
    # Strip Qwen <think> trailer if present
    if "</think>" in s:
        s = s[s.rfind("</think>") + len("</think>"):].lstrip()
    # Try direct parse
    try:
        v = json.loads(s)
        if isinstance(v, dict):
            return v
    except json.JSONDecodeError:
        pass
    # Fallback: find first/last brace
    m2 = _JSON_OBJECT_RE.search(s)
    if m2:
        try:
            v = json.loads(m2.group(0))
            if isinstance(v, dict):
                return v
        except json.JSONDecodeError:
            pass
    return None


def _validate_against_schema(
    extracted: Dict[str, Any], schema: Dict[str, Any]
) -> Tuple[Dict[str, Any], List[str]]:
    """Light schema check: ensure every schema key is present (fill
    missing with None). Returns (normalized_dict, missing_keys_list).
    No type-coercion — caller can use string values as-is."""
    if not schema:
        return extracted, []
    missing: List[str] = []
    out: Dict[str, Any] = dict(extracted)
    for k in schema:
        if k not in out:
            out[k] = None
            missing.append(k)
    return out, missing


def run_extract_info(
    *, controller, call_llm, model: str,
    paths: List[str], query: str,
    output_schema: Optional[Dict[str, Any]] = None,
    per_file_max_tokens: int = 800,
    logger: Optional[Any] = None,
) -> Dict[str, Any]:
    """Main entry point. Sequentially process each path; call planner
    LLM once per file. Returns an aggregate dict the agent persists.

    Caller is the agent — it already has ``call_llm`` (the planner LLM
    function) and ``env.controller`` for VM file access. Stays sync.

    ``logger``: optional override. When provided (typical), uses the
    caller's logger (the agent's mm_agents logger that runtime.log is
    wired to). When None, falls back to the module-level logger — which
    in the agent process has no handlers, so per-file progress
    silently disappears. THAT was the bug in v1: handle() got the
    caller logger but run_extract_info() used the unconfigured module
    one, so only the wrapping "running"/"done" lines appeared.
    """
    import time as _time
    _log = logger if logger is not None else globals()['logger']
    paths = paths[:MAX_FILES_PER_CALL]
    schema = dict(output_schema or {})
    result: Dict[str, Any] = {
        "ok": True,
        "query": query,
        "schema": schema,
        "files": [],
    }
    n_total = len(paths)
    t_overall = _time.time()
    for idx, path in enumerate(paths, start=1):
        ext = _ext_of(path)
        file_rec: Dict[str, Any] = {
            "path": path, "ext": ext, "ok": False,
            "extracted": None, "error": None, "raw_excerpt": "",
        }
        t_file = _time.time()
        _log.info("[extract_info] (%d/%d) %s — fetching", idx, n_total, path)
        data = _fetch_file(controller, path)
        if data is None:
            file_rec["error"] = "file unreadable from VM"
            result["files"].append(file_rec)
            _log.info("[extract_info] (%d/%d) %s — FAILED: file unreadable",
                        idx, n_total, path)
            continue
        _log.info("[extract_info] (%d/%d) %s — fetched %d bytes "
                    "(ext=%s), preparing LLM call",
                    idx, n_total, os.path.basename(path), len(data), ext)

        text_content: Optional[str] = None
        image_bytes: Optional[bytes] = None
        image_mime = "image/jpeg"

        if ext in _IMAGE_EXTS:
            image_bytes = _maybe_downscale_image(data, ext)
            image_mime = {
                ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
                ".png": "image/png", ".gif": "image/gif",
                ".bmp": "image/bmp", ".webp": "image/webp",
            }.get(ext, "image/jpeg")
        elif ext in _PDF_EXTS:
            txt, img = _extract_pdf_text(data)
            text_content = txt
            image_bytes = img
        elif ext in _DOCX_EXTS:
            text_content = _extract_docx_text(data)
        elif ext in _XLSX_EXTS:
            text_content = _extract_xlsx_summary(data)
        elif ext in _OFFICE_TEXT_EXTS or ext == "":
            try:
                text_content = data.decode("utf-8", errors="replace")
            except Exception:
                text_content = None
        else:
            # Try text decode as a last resort
            try:
                text_content = data.decode("utf-8", errors="replace")
            except Exception:
                file_rec["error"] = f"unsupported extension {ext!r}"
                result["files"].append(file_rec)
                continue

        if text_content is None and image_bytes is None:
            file_rec["error"] = "no extractable content (text+image both empty)"
            result["files"].append(file_rec)
            continue

        file_rec["raw_excerpt"] = (
            (text_content or "")[:RAW_EXCERPT_CHARS]
            if text_content is not None
            else f"<image, {len(image_bytes)} bytes>"
        )

        messages = _build_llm_call(
            query=query, schema=schema, path=path,
            text_content=text_content, image_bytes=image_bytes,
            image_mime=image_mime,
        )
        t_llm = _time.time()
        try:
            raw = call_llm(
                {"model": model, "messages": messages,
                 "max_tokens": per_file_max_tokens,
                 "temperature": 0.0, "top_p": 0.9},
                model,
            )
        except Exception as e:
            file_rec["error"] = f"LLM call failed: {e!r}"
            result["files"].append(file_rec)
            _log.warning("[extract_info] (%d/%d) %s — LLM call CRASHED "
                           "after %.1fs: %s",
                           idx, n_total, os.path.basename(path),
                           _time.time() - t_llm, e)
            continue
        llm_dt = _time.time() - t_llm
        if not raw:
            file_rec["error"] = "LLM returned empty"
            result["files"].append(file_rec)
            _log.warning("[extract_info] (%d/%d) %s — LLM empty after %.1fs",
                           idx, n_total, os.path.basename(path), llm_dt)
            continue

        parsed = _parse_json_response(raw)
        if parsed is None:
            file_rec["error"] = "LLM response not JSON-parseable"
            file_rec["raw_excerpt"] = (raw or "")[:RAW_EXCERPT_CHARS]
            result["files"].append(file_rec)
            _log.warning("[extract_info] (%d/%d) %s — JSON parse FAILED "
                           "(LLM %.1fs); raw head: %r",
                           idx, n_total, os.path.basename(path),
                           llm_dt, (raw or "")[:120])
            continue
        if parsed.get("_skip"):
            file_rec["ok"] = True
            file_rec["extracted"] = {"_skip": True,
                                      "reason": parsed.get("reason", "")}
            result["files"].append(file_rec)
            _log.info("[extract_info] (%d/%d) %s — SKIPPED by LLM "
                        "(LLM %.1fs): %s",
                        idx, n_total, os.path.basename(path),
                        llm_dt, parsed.get("reason", ""))
            continue
        norm, missing = _validate_against_schema(parsed, schema)
        file_rec["ok"] = True
        file_rec["extracted"] = norm
        if missing:
            file_rec["error"] = f"schema keys missing (filled with null): {missing}"
        result["files"].append(file_rec)
        # Trim values for log readability
        _summary = {
            k: (str(v)[:60] + "…" if v is not None and len(str(v)) > 60 else v)
            for k, v in norm.items()
        }
        _log.info("[extract_info] (%d/%d) %s — OK (LLM %.1fs, file %.1fs) "
                    "extracted=%s%s",
                    idx, n_total, os.path.basename(path),
                    llm_dt, _time.time() - t_file,
                    _summary,
                    f" [missing: {missing}]" if missing else "")

    # Overall ok = any file extracted successfully
    result["ok"] = any(f["ok"] for f in result["files"])
    _log.info("[extract_info] ALL DONE — %d/%d files ok, total %.1fs",
                sum(1 for f in result["files"] if f.get("ok")),
                n_total, _time.time() - t_overall)
    return result


def handle(
    *, env, action_inputs: Dict[str, Any],
    call_llm, model: str, logger: Optional[Any] = None,
) -> Dict[str, Any]:
    """Decode the decomposer-wire ``action_inputs`` and run the
    extraction. This is the single entry point the agent's actor
    pipeline should call when it sees ``action_type == "extract_info"``.

    Pulls ``paths`` / ``query`` / ``output_schema`` from action_inputs
    (the decomposer JSON-encoded them so they survive repr + literal_eval
    round-trip on the way through). Returns the aggregate dict that the
    agent persists on ``self._last_extract_output`` for the next
    planner prompt.
    """
    ctrl = getattr(env, "controller", None) if env is not None else None
    if ctrl is None or not hasattr(ctrl, "get_file"):
        return {"ok": False,
                "error": "no env.controller.get_file available"}

    # JSON-decode each wire field — defensive, accepts either str or
    # already-decoded value.
    paths_raw = action_inputs.get("paths") or "[]"
    try:
        paths = (json.loads(paths_raw) if isinstance(paths_raw, str)
                 else list(paths_raw))
        if not isinstance(paths, list):
            paths = [str(paths)]
    except Exception:
        paths = [str(paths_raw)]
    paths = [str(p) for p in paths if p]

    query = str(action_inputs.get("query") or "").strip()

    schema_raw = action_inputs.get("output_schema") \
        or action_inputs.get("schema") or "{}"
    try:
        schema = (json.loads(schema_raw) if isinstance(schema_raw, str)
                  else dict(schema_raw))
        if not isinstance(schema, dict):
            schema = {}
    except Exception:
        schema = {}

    if not paths or not query:
        return {"ok": False,
                "error": f"missing paths or query (paths={paths}, query={query!r})"}

    if logger:
        logger.info("[extract_info] running on %d file(s) with query=%r "
                    "schema_keys=%s", len(paths), query[:80], list(schema))
    try:
        result = run_extract_info(
            controller=ctrl, call_llm=call_llm, model=model,
            paths=paths, query=query, output_schema=schema,
            logger=logger,   # thread the caller's logger so per-file
                              # progress lands in runtime.log
        )
    except Exception as e:
        if logger:
            logger.warning("[extract_info] crashed: %s", e)
        return {"ok": False, "error": f"run_extract_info crashed: {e!r}"}

    if logger:
        n_ok = sum(1 for f in result.get("files", []) if f.get("ok"))
        logger.info("[extract_info] done: %d/%d files ok",
                    n_ok, len(result.get("files", [])))
    return result


def render_for_planner(
    result: Optional[Any],
    max_chars: int = 8000,
) -> Optional[str]:
    """Render extract_info result(s) into a planner-prompt block.

    Accepts either a single batch dict (legacy shape) OR a list of batch
    dicts (multi-step extraction accumulator). Returns None when there's
    nothing to show. Compact tabular form — one row per file, with
    extracted JSON inline. Per-file rows are deduped by path within each
    batch keep the latest occurrence; across batches the same path may
    appear more than once (the planner sees both attempts so it can pick
    the better extraction). When rendering a list, batches are labeled
    so the planner can track which call produced which rows.
    """
    if not result:
        return None
    # Normalize to list-of-batches for unified rendering.
    if isinstance(result, dict):
        batches: List[Dict[str, Any]] = [result]
    elif isinstance(result, list):
        batches = [b for b in result if isinstance(b, dict)]
    else:
        return None
    # Drop empty batches.
    batches = [b for b in batches if (b.get("files") or [])]
    if not batches:
        return None

    lines: List[str] = []
    lines.append("[extract_info results — agent already read these files; "
                 "do NOT plan GUI steps to re-open them. The full history "
                 "across this task's extract_info calls is shown below; "
                 "use ALL of it when planning downstream writes.]")
    multi = len(batches) > 1
    for i, batch in enumerate(batches, 1):
        files = batch.get("files") or []
        if multi:
            lines.append("")
            lines.append(f"── Batch {i}/{len(batches)} ──")
        lines.append(f"Query: {batch.get('query', '')[:200]}")
        schema = batch.get("schema") or {}
        if schema:
            lines.append(f"Schema: {json.dumps(schema)}")
        lines.append("")
        for f in files:
            path = f.get("path", "")
            ok = "✓" if f.get("ok") else "✗"
            if f.get("ok"):
                extracted = f.get("extracted") or {}
                lines.append(f"  {ok} {path}")
                lines.append(f"      {json.dumps(extracted, ensure_ascii=False)}")
            else:
                err = f.get("error") or "?"
                lines.append(f"  {ok} {path}   ERROR: {err}")
    rendered = "\n".join(lines)
    if len(rendered) > max_chars:
        rendered = rendered[:max_chars] + "\n  …[truncated]"
    return rendered
